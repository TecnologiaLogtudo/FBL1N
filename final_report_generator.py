# final_report_generator.py
import pandas as pd
import numpy as np
from utils import logger, format_currency

class FinalReportGenerator:
    """
    Classe responsável por gerar o arquivo Excel final com as análises
    e tabelas dinâmicas a partir dos dados já processados.
    """
    def __init__(self, analyzed_df):
        """
        Inicializa o gerador com o DataFrame final analisado.
        
        Args:
            analyzed_df (pd.DataFrame): O DataFrame vindo do AnalysisProcessor.
        """
        self.analyzed_df = analyzed_df.copy() if analyzed_df is not None else pd.DataFrame()
        logger.info("Gerador de Relatório Final inicializado.")

    def _format_currency_column(self, series):
        """Formata uma série de valores numéricos como moeda brasileira."""
        return series.apply(lambda x: format_currency(x) if pd.notna(x) and x != 0 else np.nan)

    def _apply_excel_styling(self, worksheet, dataframe, table_start_row=3):
        """
        Aplica formatação executiva ao worksheet do Excel.
        
        Args:
            worksheet: Objeto worksheet do ExcelWriter.
            dataframe: DataFrame que foi escrito na planilha.
            table_start_row: A linha onde o cabeçalho da tabela começa (1-indexed).
        """
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # --- Definição de Estilos ---
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        total_font = Font(bold=True)
        total_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        
        # Cor de fundo laranja para Logtudo Ceará
        ceara_fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')

        # Definição de Bordas
        thin_side = Side(style='thin')
        thick_side = Side(style='thick')
        
        # --- Largura das Colunas ---
        column_widths = {'A': 25, 'B': 20, 'C': 20, 'D': 20, 'E': 20}
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # --- Formatação do Cabeçalho (Subtítulos) ---
        header_cells = worksheet[table_start_row]
        for cell in header_cells:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # --- Formatação das Linhas de Dados ---
        data_start_row = table_start_row + 1
        num_data_rows = len(dataframe)
        
        # Itera sobre as linhas, mas limita a iteração de colunas de A a E (1 a 5)
        for r_idx, sheet_row in enumerate(worksheet.iter_rows(min_row=data_start_row, max_row=data_start_row + num_data_rows - 1, min_col=1, max_col=5)):
            df_row = dataframe.iloc[r_idx]
            is_total_row = df_row['Serviço'] == 'Total Geral'
            is_ceara_row = df_row['Transportadora'] == 'Logtudo Ceará'

            for cell in sheet_row:
                # Aplica fundo laranja para linhas da Logtudo Ceará (exceto total)
                if is_ceara_row and not is_total_row:
                    cell.fill = ceara_fill

                # Aplica estilo para linhas de "Total Geral"
                if is_total_row:
                    cell.font = total_font
                    cell.fill = total_fill

        # --- Aplicação de Bordas ---
        num_cols = len(dataframe.columns)
        max_table_row = table_start_row + num_data_rows
        
        for r_idx, row in enumerate(worksheet.iter_rows(min_row=table_start_row, max_row=max_table_row, min_col=1, max_col=num_cols)):
            df_row_index = r_idx - 1 # Index para o DataFrame (-1 para header)

            for cell in row:
                top = thin_side
                left = thin_side
                right = thin_side
                bottom = thin_side

                # Bordas externas grossas
                if cell.row == table_start_row: top = thick_side
                if cell.column == 1: left = thick_side
                if cell.column == num_cols: right = thick_side
                if cell.row == max_table_row: bottom = thick_side
                
                # Bordas grossas entre transportadoras
                if df_row_index >= 0 and df_row_index < num_data_rows - 1:
                    current_transportadora = dataframe.iloc[df_row_index]['Transportadora']
                    next_transportadora = dataframe.iloc[df_row_index + 1]['Transportadora']
                    
                    if current_transportadora != next_transportadora and current_transportadora.startswith('Logtudo'):
                        bottom = thick_side

                cell.border = Border(top=top, left=left, right=right, bottom=bottom)

    def _apply_detail_table_styling(self, worksheet, dataframe, table_start_row=3, table_start_col=7):
        """
        (NOVO) Aplica formatação executiva à tabela de detalhes na aba de resumo.

        Args:
            worksheet: Objeto worksheet do ExcelWriter.
            dataframe: DataFrame da tabela de detalhes.
            table_start_row: Linha onde o cabeçalho da tabela começa (1-indexed).
            table_start_col: Coluna onde a tabela começa (1-indexed).
        """
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        logger.debug(f"Aplicando estilo na tabela de detalhes a partir da linha {table_start_row}, coluna {table_start_col}.")

        # --- Definição de Estilos ---
        header_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        # Uma cor de cabeçalho diferente para distinguir da tabela principal
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(horizontal='left', vertical='center')
        
        thin_side = Side(style='thin', color='BFBFBF')
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        num_cols = len(dataframe.columns)
        num_rows = len(dataframe)

        # --- Largura das Colunas ---
        # Mapeia o índice da coluna (a partir de 0) para a largura desejada
        column_widths = [12, 12, 22, 12, 28, 18, 15, 15, 20, 15, 15, 15, 15, 15]
        
        for i, width in enumerate(column_widths):
            col_letter = get_column_letter(table_start_col + i)
            worksheet.column_dimensions[col_letter].width = width
            logger.debug(f"Definindo largura da coluna {col_letter} para {width}.")

        # --- Formatação do Cabeçalho ---
        for col_idx in range(num_cols):
            cell = worksheet.cell(row=table_start_row, column=table_start_col + col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        logger.debug("Cabeçalho da tabela de detalhes formatado.")

        # --- Formatação das Células de Dados ---
        for row_idx in range(num_rows):
            for col_idx in range(num_cols):
                cell = worksheet.cell(row=table_start_row + 1 + row_idx, column=table_start_col + col_idx)
                cell.border = border
                cell.alignment = cell_alignment
                
                # Formatação especial para colunas de valor
                col_name = dataframe.columns[col_idx]
                if col_name in ['Valor CTe', 'Valor pago', 'Recebido/A receber']:
                    cell.number_format = 'R$ #,##0.00'
        
        logger.debug(f"{num_rows} linhas de dados formatadas na tabela de detalhes.")


    def _generate_comprehensive_summary(self):
        """
        Gera uma tabela única consolidada com todas as transportadoras e serviços.
        """
        logger.info("Gerando tabela consolidada de resumo...")
        
        if self.analyzed_df.empty:
            logger.warning("DataFrame de análise está vazio. Nenhuma tabela será gerada.")
            return pd.DataFrame()

        self.analyzed_df['Recebido/A receber'] = pd.to_numeric(self.analyzed_df['Recebido/A receber'], errors='coerce').fillna(0)
        
        expected_services = [
            'Complemento', 'Descarga', 'Diária no cliente', 'Diária parado', 
            'Frete', 'Pedágio', 'Reentrega'
        ]
        
        transportadoras = ['Logtudo Bahia', 'Logtudo Ceará', 'Logtudo Pernambuco']
        
        all_results = {}
        
        for transportadora in transportadoras:
            logger.info(f"Processando transportadora: {transportadora}")
            
            df_transportadora = self.analyzed_df[
                (self.analyzed_df['Transportadora'] == transportadora) &
                (self.analyzed_df['Serviço'].isin(expected_services))
            ].copy()
            
            pivot = pd.pivot_table(
                df_transportadora,
                values='Recebido/A receber',
                index='Serviço',
                columns='Status pgto',
                aggfunc='sum',
                fill_value=0,
                margins=True,
                margins_name='Total Geral'
            ).rename_axis(None, axis=1).reset_index()
            
            for status in ['Não compensado', 'Não lançado']:
                if status not in pivot.columns:
                    pivot[status] = 0
            
            for service in expected_services:
                if service not in pivot['Serviço'].values:
                    new_row = pd.DataFrame({'Serviço': [service], 'Não compensado': [0], 'Não lançado': [0]})
                    if 'Total Geral' in pivot.columns:
                        new_row['Total Geral'] = 0
                    pivot = pd.concat([pivot, new_row], ignore_index=True)
            
            pivot = pivot.set_index('Serviço').reindex(expected_services + ['Total Geral']).reset_index()
            
            for i in range(len(pivot) - 1):
                pivot.loc[i, 'Total Geral'] = pivot.loc[i, 'Não compensado'] + pivot.loc[i, 'Não lançado']
            
            pivot.loc[pivot['Serviço'] == 'Total Geral', 'Total Geral'] = \
                pivot.loc[pivot['Serviço'] == 'Total Geral', 'Não compensado'] + \
                pivot.loc[pivot['Serviço'] == 'Total Geral', 'Não lançado']
            
            for col in ['Não compensado', 'Não lançado', 'Total Geral']:
                pivot[col] = self._format_currency_column(pivot[col])
            
            transportadora_results = {}
            for _, row in pivot.iterrows():
                service_name = row['Serviço']
                transportadora_results[service_name] = {
                    'Não compensado': row['Não compensado'],
                    'Não lançado': row['Não lançado'],
                    'Total Geral': row['Total Geral']
                }
            
            all_results[transportadora] = {'Transporte/Serviço': transportadora_results}
        
        logger.info("Tabela consolidada gerada com sucesso.")
        return all_results

    def _create_dataframe_from_nested_dict(self, nested_dict):
        """Converte o dicionário aninhado em um DataFrame para exportação."""
        data_rows = []
        
        # Garante a ordem correta das transportadoras
        transportadoras_ordered = ['Logtudo Bahia', 'Logtudo Ceará', 'Logtudo Pernambuco']
        services_ordered = ['Complemento', 'Descarga', 'Diária no cliente', 'Diária parado', 'Frete', 'Pedágio', 'Reentrega']

        for transportadora in transportadoras_ordered:
            if transportadora in nested_dict:
                services_data = nested_dict[transportadora]
                for service_name in services_ordered:
                     if service_name in services_data['Transporte/Serviço']:
                        metrics = services_data['Transporte/Serviço'][service_name]
                        row = {
                            'Transportadora': transportadora,
                            'Serviço': service_name,
                            'Não compensado': metrics['Não compensado'],
                            'Não lançado': metrics['Não lançado'],
                            'Total Geral': metrics['Total Geral']
                        }
                        data_rows.append(row)
        
        return pd.DataFrame(data_rows)

    def generate_report(self, writer):
        """
        Gera o relatório final consolidado com formatação executiva.
        """
        logger.info("--- INICIANDO GERAÇÃO DO RELATÓRIO FINAL CONSOLIDADO ---")
        
        consolidated_data = self._generate_comprehensive_summary()
        export_df = self._create_dataframe_from_nested_dict(consolidated_data)
        
        totals_by_transportadora = []
        transportadoras_list = ['Logtudo Bahia', 'Logtudo Ceará', 'Logtudo Pernambuco']
        for transportadora in transportadoras_list:
            transport_data = export_df[export_df['Transportadora'] == transportadora]
            
            total_nao_compensado = 0
            total_nao_lancado = 0
            
            for value in transport_data['Não compensado']:
                if pd.notna(value) and value != 'nan':
                    try:
                        clean_value = float(str(value).replace('R$', '').replace('.', '').replace(',', '.'))
                        total_nao_compensado += clean_value
                    except (ValueError, TypeError): pass
            
            for value in transport_data['Não lançado']:
                if pd.notna(value) and value != 'nan':
                    try:
                        clean_value = float(str(value).replace('R$', '').replace('.', '').replace(',', '.'))
                        total_nao_lancado += clean_value
                    except (ValueError, TypeError): pass
            
            total_geral = total_nao_compensado + total_nao_lancado
            
            # Para a linha de total, a coluna "Transportadora" conterá o nome da transportadora
            totals_by_transportadora.append({
                'Transportadora': transportadora,
                'Serviço': 'Total Geral',
                'Não compensado': format_currency(total_nao_compensado),
                'Não lançado': format_currency(total_nao_lancado),
                'Total Geral': format_currency(total_geral)
            })
        
        totals_df = pd.DataFrame(totals_by_transportadora)
        final_export_df = pd.concat([export_df, totals_df], ignore_index=True)
        
        # --- Escrita no Excel ---
        sheet_name = 'Resumo Consolidado'
        # Deixa uma linha em branco para o título e outra para espaçamento
        final_export_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        
        worksheet = writer.sheets[sheet_name]
        
        # --- Título Executivo (acima da tabela) ---
        from openpyxl.styles import Font, Alignment
        
        title_font = Font(bold=True, size=16, color='000000')
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        # (ALTERADO) Ajusta o merge para cobrir ambas as tabelas (A até T)
        worksheet.merge_cells('A1:T1')
        title_cell = worksheet.cell(row=1, column=1, value='RESUMO CONSOLIDADO DE TRANSPORTES')
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        
        # --- Aplica todos os outros estilos ---
        self._apply_excel_styling(worksheet, final_export_df, table_start_row=3)
        
        # --- (NOVO) GERAÇÃO E ADIÇÃO DA TABELA DE DETALHES ---
        logger.info("Iniciando a criação da tabela de detalhes de pagamentos pendentes.")

        # 1. Filtra o DataFrame original com os status desejados
        status_filter = ['Não lançado', 'Não compensado']
        filtered_details_df = self.analyzed_df[self.analyzed_df['Status pgto'].isin(status_filter)].copy()
        
        if filtered_details_df.empty:
            logger.warning("Nenhuma linha encontrada com status 'Não lançado' ou 'Não compensado'. A tabela de detalhes não será adicionada.")
        else:
            logger.info(f"Foram encontradas {len(filtered_details_df)} linhas para a tabela de detalhes.")

            # 2. Seleciona e reordena as colunas para a tabela de detalhes
            columns_to_keep = [
                'Emissao', 'Mês', 'Transportadora', 'CTRC', 'Cliente', 'Serviço',
                'Senha Ravex', 'DT Frete', 'Destino', 'Nota Fiscal', 'Valor CTe',
                'Status pgto', 'Valor pago', 'Recebido/A receber'
            ]
            
            final_details_df = filtered_details_df[columns_to_keep]

            # Renomeia colunas para uma apresentação mais clara
            final_details_df = final_details_df.rename(columns={
                'Emissao': 'Emissão',
                'Nota Fiscal': 'Nota fiscal',
                'Status pgto': 'Status Pgto'
            })
            
            # Converte as colunas de valor para numérico para formatação correta no Excel
            for col in ['Valor CTe', 'Valor pago', 'Recebido/A receber']:
                if col in final_details_df.columns:
                    final_details_df[col] = pd.to_numeric(final_details_df[col], errors='coerce')

            logger.info("Colunas para a tabela de detalhes foram selecionadas e renomeadas.")

            # Define a coluna inicial para a nova tabela (G = 7, mas pandas usa índice 0, então 6)
            start_col_details = 6 
            
            # 3. Escreve a tabela de detalhes na planilha
            final_details_df.to_excel(
                writer, 
                sheet_name=sheet_name, 
                index=False, 
                startrow=2, 
                startcol=start_col_details
            )
            logger.info(f"Tabela de detalhes escrita na aba '{sheet_name}' a partir da coluna G.")
            
            # 4. Aplica a formatação executiva na nova tabela
            # A coluna inicial para openpyxl é 1-indexed (G=7)
            self._apply_detail_table_styling(worksheet, final_details_df, table_start_row=3, table_start_col=start_col_details + 1)
            logger.info("Formatação executiva aplicada à tabela de detalhes.")
        # --- FIM DA ADIÇÃO DA TABELA DE DETALHES ---

        final_report_sheets = {sheet_name: final_export_df}
        
        logger.info("--- GERAÇÃO DO RELATÓRIO FINAL CONCLUÍDA. ---")
        return final_report_sheets
